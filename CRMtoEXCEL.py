from turtle import position
import pyautogui as gui
import subprocess
import time
import glob
import pathlib
import win32gui
import numpy as np 
import pandas as pd
import win32con
import os
import argparse
import openpyxl as px
from openpyxl.styles import Border, Side, PatternFill, Font
import pyperclip
import shutil
from playsound import playsound
import random
import datetime

#一度の最大処理ファイル数
size = 30

commencer = time.time()

#座標----------------------------------------------------------------------------------------------------------------
positionOpenFile = (28, 65)#
positionChannel2 = (393,595)#
positionFile = (36,36) #
positionContinuousExport = (75,207)#
positionOverwrite = (67,132)#
positionOutRange = (539, 1)#
positionPeakButton = (401,68)#
positionBaseBelowButton = (279,65)#
positionBaseAboveButton = (312,63)#
isPath = False
#座標----------------------------------------------------------------------------------------------------------------


#CMRProを最前面に
def CRMfront():
    try:
        memoapp = win32gui.FindWindow(None,"Chromato-PRO [波形解析]")
        time.sleep(3)
        win32gui.SetForegroundWindow(memoapp)
    except:
        time.sleep(5)
        memoapp = win32gui.FindWindow(None,"Chromato-PRO [波形解析]")
        time.sleep(3)
        win32gui.SetForegroundWindow(memoapp)

#ファイルを開くを押す
def ClickOpenFile():
    gui.click(positionOpenFile)

#ファイル名入力
def WriteFileName(name):
    pyperclip.copy(file_ab)
    time.sleep(0.5)
    gui.hotkey("ctrl","v")
    time.sleep(1)
    gui.press("enter")

#チャネル変更(左下をクリック)
def Chanel2():
    gui.click(positionChannel2)

#エクセルに連続エクスポート
def ExcelExport():
    gui.click(positionFile)
    gui.click(positionContinuousExport)
    time.sleep(1)
    #エクスポートボタンを認識してクリック
    path = r".\images\exportButton.png"
    left, top, width, height = gui.locateOnScreen(path, confidence = 0.6, grayscale = True)
    gui.click(left+width//4, top+height//2)

#エクセルファイルを前面に
def ExcelFront():
    try:
        memoapp = win32gui.FindWindow(None,"Book1 - Excel")
        time.sleep(4)
        win32gui.SetForegroundWindow(memoapp)
    except Exception as e:
        time.sleep(5)
        memoapp = win32gui.FindWindow(None,"Book1 - Excel")
        time.sleep(4)
        win32gui.SetForegroundWindow(memoapp)
    

#ファイルを保存
def TempSave(name):
    gui.hotkey("ctrl","s")
    time.sleep(1.5)
    gui.write(name)
    time.sleep(2)
    gui.press("enter")

def exportcheck(nom):
    wb = px.load_workbook(f'{exname}.xlsx')
    ws = wb[nom]
    #ファイル最大行の取得
    maxr = ws.max_row  
    sheet_range = ws["A2":f"A{maxr}"]
    lst = []
    #A列のA2から最後まで取得
    for row in sheet_range:
        for cell in row:
    # 該当セルの値取得
            cell_value = cell.value
            lst.append(cell_value)
    Miss = []
    for file in all_files:
        if file not in lst:
            Miss.append(file)
    return Miss

def excelarrange(nom):
    #ファイルを読み込んでシートを選択
    wb = px.load_workbook(f'{exname}.xlsx')
    ws = wb[nom]
    #罫線の作成
    side = Side(style = 'thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    #罫線作成
    for rows in ws:
        for cell in rows:
            cell.border = border
    #文字の色を変更
    # 背景色を変更
    maxr = ws.max_row
    for rows in ws['A1':'E1']:
        for cell in rows:
            cell.fill = PatternFill(patternType='solid',fgColor='D1FE7B',bgColor='D1FE7B')
            cell.font = Font(bold=True)
    for rows in ws["A2":f"A{maxr}"]:
        for cell in rows:
            cell.font = Font(bold = False)
    #列幅調節
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        columnstr = px.utils.get_column_letter(column)
    
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[columnstr].width = adjusted_width
    wb.save(f'{exname}.xlsx')

def copyfiles_start():
    path_CRMtoEXCEL = os.getcwd()
    files = glob.glob(r"{}/*.crm".format(input_path))
    for file in files:
        shutil.copy(file,f"{path_CRMtoEXCEL}")

def movefiles_end():
    path_CRMtoEXCEL = os.getcwd()
    files = glob.glob(r"{}/*.xlsx".format(path_CRMtoEXCEL))
    os.makedirs(output_path, exist_ok=True) 
    for file in files:
        try:
            shutil.move(file,f"{output_path}")
        except:
            os.remove(f"{output_path}/{os.path.basename(file)}")
            shutil.move(file,f"{output_path}")

#入力ボックスが消えてピークチェックできる状態になっているか確認
def inputCheck():
    imagePath = r"./images/inputButton.png"
    check = gui.locateOnScreen(imagePath, confidence = 0.7, grayscale = True)
    #入力ボタンがあるとTrueで返すので反転
    return not check

def pathCheck():
    global isPath
    if isPath:
        return
    imagePath = r"./images/checkPath.png"
    check = gui.locateOnScreen(imagePath, confidence = 0.9, grayscale = True)
    if not check:
        raise NameError("ファイルを開く -> ファイルの場所が「CRMProAuto」じゃないのでは??")
    else:
        isPath = True

parser = argparse.ArgumentParser()
parser.add_argument("--path",action = "store_true", help="CRMtoEXCELフォルダ以外のファイルを実行したい時に使用して下さい。")
args = parser.parse_args()

#音声データ
startsound = glob.glob("./sound/start*")
endsound = glob.glob("./sound/end*")

#入力
exname = input("出力後エクセルファイル名を入力して下さい:")
isChannel_2_bool = input("Channel_1 のみ処理しますか? [y/n]:")

#Channel 1のみ処理するか否か
if isChannel_2_bool == "y":
    isChannel_2 = False
    print("Channel_1 のみ処理します。")
else:
    isChannel_2 = True


file_list=glob.glob("C:\\Users\\四国めたん\\Documents\\CRMtoEXCELtemp*.xlsx")
for i in file_list:
    os.remove(i)

#フォルダの掃除
if args.path == True:
    pre_files = glob.glob("*.crm") + glob.glob("*.xlsx")
    if bool(pre_files) == True:
        try:
            os.mkdir("./previous")
        except:
            pass
        for file in pre_files:
            try:
                shutil.move(f"{file}","./previous")
            except shutil.Error:
                os.remove(file)

#パスの確認
if args.path == True:
    input_path = input("処理する.crmファイルが含まれているフォルダの絶対パスを入力して下さい。:").replace('"','')
    output_path = input("出力エクセルファイルを保存するフォルダの絶対パスを入力して下さい。:").replace('"','')
    if not os.path.exists(input_path):
        raise Exception("処理するフォルダが見当たりません")
    if not os.path.exists(output_path):
        raise Exception("出力するフォルダが見当たりません")
    copyfiles_start()


soundsta = random.randrange(len(startsound))
playsound(startsound[soundsta])

place = str(os.getcwd())

#Channel_1のエクスポート#

#フォルダ内のcrmファイルを抽出
all_files = glob.glob("*.crm")

#繰り返し数
itr = -(-len(all_files)//size)

#ファイル何個かずつで繰り返し
lst = []
for (start, count) in zip(range(0, len(all_files),size),range(1,itr+1)):
    files = all_files[start:start+size]

#アプリ起動
    time.sleep(5)
    sup = subprocess.Popen(r'C:\Program Files (x86)\runtime\Chromato-PRO\ChromatoPro.exe')
    time.sleep(5)

#CMRProを最前面に
    CRMfront()
    time.sleep(2)

#全画面
    gui.hotkey("win","up")
    time.sleep(5)

#crmファイルの繰り返し処理
    for file in files:

        #ファイルのパス取得
        file_p = pathlib.Path(file)
        file_ab = str(file_p)

#ファイルを開くを押す
        ClickOpenFile()
        time.sleep(3)
        pathCheck()

#ファイル名入力
        WriteFileName(file_ab)
        time.sleep(3)

#エクセルに連続エクスポート
    ExcelExport()

    time.sleep(5)
#エクセルファイルを前面に
    ExcelFront()
    time.sleep(4)
#ファイルを保存
    TempSave(f"CRMtoEXCELtemp{count}")
#クロマトを閉じる
    sup.kill()

time.sleep(5)

#分割したエクセルを纏める
# ファイルリスト取得
file_list=glob.glob("C:\\Users\\四国めたん\\Documents\\CRMtoEXCELtemp*.xlsx")
# pandasのデータフレーム化
df = pd.DataFrame()
for file in file_list:

#エクセルの読み込み
    df2 = pd.read_excel(file, parse_dates=True,index_col=0)

#欠損値を埋める
    df2 = df2.replace(np.nan,' ', regex=True)

#ファイルを結合
    df = pd.concat([df, df2])

# 一つにまとめたエクセルファイルを保存
df.to_excel(r"{}\{}.xlsx".format(place,exname),sheet_name='Channel_1')

#エクセルの終了と一時保存ファイルの削除
for i in range(1,itr+1):
    try:
        handle = win32gui.FindWindow(None,f"CRMtoEXCELtemp{i}.xlsx - Excel")
        time.sleep(3)
        win32gui.PostMessage(handle,win32con.WM_CLOSE,0,0)
    except:
        time.sleep(5)
        handle = win32gui.FindWindow(None,f"CRMtoEXCELtemp{i}.xlsx - Excel")
        time.sleep(3)
        win32gui.PostMessage(handle,win32con.WM_CLOSE,0,0)
    time.sleep(3)
    os.remove(f"C:\\Users\\四国めたん\\Documents\\CRMtoEXCELtemp{i}.xlsx")

#CH2のエクスポート#
if isChannel_2 == True:
    #ファイル何個かずつで繰り返し
    lst = []
    for (start, count) in zip(range(0, len(all_files),size),range(1,itr+1)):
        files = all_files[start:start+size]
        

    #アプリ起動
        time.sleep(5)
        sup = subprocess.Popen(r'C:\Program Files (x86)\runtime\Chromato-PRO\ChromatoPro.exe')
        time.sleep(5)

    #CMRProを最前面に
        CRMfront()
        time.sleep(2)

    #全画面
        gui.hotkey("win","up")
        time.sleep(5)

    #crmファイルの繰り返し処理
        for file in files:

            #ファイルのパス取得
            file_p = pathlib.Path(file)
            file_ab = str(file_p)

    #ファイルを開くを押す
            ClickOpenFile()
            time.sleep(3)

    #ファイル名入力
            WriteFileName(file_ab)
            time.sleep(3)

    #チャネル変更
            Chanel2()

    #エクセルに連続エクスポート
        ExcelExport()
        time.sleep(5)

    #エクセルファイルを前面に
        ExcelFront()
        time.sleep(2)

    #ファイルを保存
        TempSave(f"CRMtoEXCELtemp{count}")
        time.sleep(4)
    #クロマトを閉じる
        sup.kill()

        time.sleep(5)

    #分割したエクセルを纏める
    # ファイルリスト取得
    file_list=glob.glob("C:\\Users\\四国めたん\\Documents\\CRMtoEXCELtemp*.xlsx")
    # pandasのデータフレーム化
    df = pd.DataFrame()
    for file in file_list:

    #エクセルの読み込み
        df2 = pd.read_excel(file, parse_dates=True,index_col=0)

    #欠損値を埋める
        df2 = df2.replace(np.nan,' ', regex=True)

    #ファイルを結合
        df = pd.concat([df, df2])

    # 一つにまとめたエクセルファイルを保存
    with pd.ExcelWriter(r"{}\{}.xlsx".format(place,exname), engine="openpyxl", mode="a") as writer:
        df.to_excel(writer,sheet_name='Channel_2')

    #エクセルの終了と一時保存ファイルの削除
    for i in range(1,itr+1):
        try:
            handle = win32gui.FindWindow(None,f"CRMtoEXCELtemp{i}.xlsx - Excel")
            time.sleep(3)
            win32gui.PostMessage(handle,win32con.WM_CLOSE,0,0)
        except Exception as e:
            time.sleep(5)
            handle = win32gui.FindWindow(None,f"CRMtoEXCELtemp{i}.xlsx - Excel")
            time.sleep(3)
            win32gui.PostMessage(handle,win32con.WM_CLOSE,0,0)
        time.sleep(3)
        os.remove(f"C:\\Users\\四国めたん\\Documents\\CRMtoEXCELtemp{i}.xlsx")

#エクセルの体裁を整える

excelarrange("Channel_1")
if isChannel_2 == True:
    excelarrange("Channel_2")


# print as 00:00:00
process_time = time.time() - commencer
elapsed_hour = process_time // 3600
elapsed_minute = (process_time - elapsed_hour*3600) // 60
elapsed_second = process_time - elapsed_hour*3600 - elapsed_minute*60

print("Complete")

#エクスポートが無事完了したかの確認
book = px.load_workbook("{}\{}.xlsx".format(place,exname))
Channel_1 = book["Channel_1"].max_row - 1
if isChannel_2 == True:
    Channel_2 = book["Channel_2"].max_row - 1
data = len(all_files)

if isChannel_2 == True:
    if Channel_1 == data and Channel_2 == data:
        print("全てのファイルをエクスポートすることができました。")
    else:
        print("以下のファイルがエクスポートできませんでした")
        print("*Channel_1*")
        print(exportcheck("Channel_1"))
        print("*Channel_2*")
        print(exportcheck("Channel_2"))
        try:
            os.mkdir("Re_excel_Channel_1")
        except:
            pass
        try:
            os.mkdir("Re_excel_Channel_2")
        except:
            pass
        for file in exportcheck("Channel_1"):
            shutil.copy(file,"./Re_excel_Channel_1")
        for file in exportcheck("Channel_2"):
            shutil.copy(file,"./Re_excel_Channel_2")

else:
    if Channel_1 == data:
        print("全てのファイルをエクスポートすることができました。")
    else:
        print("以下のファイルがエクスポートできませんでした")
        print("*Channel_1*")
        print(exportcheck("Channel_1"))
        try:
            os.mkdir("Re_excel_Channel_1")
        except:
            pass
        for file in exportcheck("Channel_1"):
            shutil.copy(file,"./Re_excel_Channel_1")

if args.path == True:
    movefiles_end()

try:
    with open(r"data/CtoE-log.csv","a") as f:
        f.write(f"{datetime.datetime.now()},{len(all_files)},{process_time}\n")

except:
    with open(r"data/CtoE-log.csv","w") as f:
        f.write(f"date, データ数 \n {datetime.datetime.now()},{len(all_files)},{process_time}\n")

soundend = random.randrange(len(endsound))
playsound(endsound[soundend])

print("Complete")
print("経過時間は{:.0f}時間{:.0f}分{:.0f}秒です".format(elapsed_hour, elapsed_minute, elapsed_second))


#エクセルを表示
if args.path == True:
    subprocess.Popen([r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE",r"{}\{}.xlsx".format(output_path,exname)])
else:
    subprocess.Popen([r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE",r"{}\{}.xlsx".format(place,exname)])
