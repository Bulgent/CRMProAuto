from config import *
from turtle import position
import numpy as np
import pandas as pd
import win32con
import openpyxl as px
from openpyxl.styles import Border, Side, PatternFill, Font
import cv2
import csv
import matplotlib.pyplot as plt
import numpy as np
import pyautogui as gui
import subprocess
import time
import win32gui
from PIL import ImageGrab, Image, ImageDraw, ImageFont
import glob
import os
import pathlib
import argparse
import pyperclip
import shutil
from playsound import playsound
import random
import datetime
import math

def CRMfront():
    '''CMRProを最前面に'''
    try:
        CRMPro_interface = win32gui.FindWindow(None, "Chromato-PRO [波形解析]")
    except:
        time.sleep(8)
        CRMPro_interface = win32gui.FindWindow(None, "Chromato-PRO [波形解析]")
    time.sleep(5)
    try:
        win32gui.SetForegroundWindow(CRMPro_interface)
    except:
        time.sleep(8)
        CRMPro_interface = win32gui.FindWindow(None, "Chromato-PRO [波形解析]")
        time.sleep(6)
        win32gui.SetForegroundWindow(CRMPro_interface)


def CRMSetting():
    '''CMRProのセッティング'''
    try:
        # アプリ起動
        CRMPro_exe = subprocess.Popen(pathCRMPro)
        time.sleep(4)
        # CMRProを最前面に
        CRMfront()
        time.sleep(2)
    except:
        # アプリ起動
        time.sleep(5)
        CRMPro_exe = subprocess.Popen(pathCRMPro)
        time.sleep(4)
        # CMRProを最前面に
        CRMfront()
        time.sleep(2)
    return CRMPro_exe  # 後でkillする


def ClickOpenFile():
    '''ファイルを開くを押す'''
    gui.click(positionOpenFile)


def WriteFileName(name):
    '''ファイル名入力
        name: 入力ファイル名
    '''
    time.sleep(0.5)
    gui.write(name)
    time.sleep(3)
    gui.press("enter")


def Save():
    '''上書き保存'''
    gui.click(positionFile)
    time.sleep(0.1)
    gui.click(positionOverwrite)


def getAllCoodinate(image, isDebug=False):
    '''グラフの全座標取得
        ・image: スクリーンショット
        ・isDebug: デバッグモードかどうか(処理過程の画像が出力される)
          return
        ・maxxc: グラフのx軸の最大値
        ・minxc: グラフのx軸の最小値
        ・cood: グラフの座標{x座標: y座標, ...}
    '''
    # 画像の読み込み
    img = cv2.imread(image, 0)
    # トリミング
    img = img[ymin:ymax, xmin:xmax]
    # 画像の範囲取得
    height, width, channels = img.shape[:3]
    # 画像の平滑化
    blurs = cv2.blur(img, (blurLevel, blurLevel))
    if isDebug:
        cv2.imwrite("blur.png", blurs)
    # グラフの端を白塗り・下に余白追加(輪郭を捉えるため)
    # 縦に白塗り(幅を足してない、x座標の変更はめんどくさいので...)
    padding = 10
    blurs = cv2.line(blurs, (0, 0), (0, height), (255, 255, 255), padding)
    blurs = cv2.line(blurs, (width, 0), (width, height),
                     (255, 255, 255), padding)
    # 横に余白(高さはあくまで比較、値そのものを使うわけではない)
    blurs = cv2.copyMakeBorder(
        blurs, 50, 50, 0, 0, cv2.BORDER_CONSTANT, value=[255, 255, 255])

    if isDebug == True:
        cv2.imwrite("blur_cut.png", blurs)
    # 画像の2極化
    retval, im_bw = cv2.threshold(blurs, 254, 255, cv2.THRESH_BINARY)
    if isDebug:
        cv2.imwrite("blur_cut.png", im_bw)
    contours, hierarchy = cv2.findContours(
        im_bw, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
    # 余計な次元(余計な輪郭)削除 (NumContours, 1, NumPoints) -> (NumContours, NumPoints)
    contours = [np.squeeze(cnt, axis=1) for cnt in contours]
    conts = contours[0]  # [[x,y],[x,y],....]
    # x座標範囲確認
    maxxc = 0
    minxc = 1000
    for cont in conts:
        if cont[0] >= maxxc:
            maxxc = cont[0]
        if cont[0] <= minxc:
            minxc = cont[0]
    rangecheck = True
    # 輪郭が取得できてるか確認(できてなければ画像加工後リトライ)
    if maxxc != width-padding//2 and minxc != padding//2:
        blurs = cv2.blur(img, (blurLevel, blurLevel))
        cv2.imwrite("autopeak_blur_re.png", blurs)
        # time.sleep(0.1)
        # blurs = cv2.imread("autopeak_blur_re.png")
        rangecheck = False
    # 足りてない部分に黒点追加して補完(補完箇所はグラフのベースライン, つまりピークで切れている場合には対応していない?)
        checkx = [False]*width
        for numx in range(1, width):
            for numy in range(1, height):
                # もしそのx座標のどこかのピクセルが白でなければTrue
                if blurs.item(numy, numx, 0) != 255 or blurs.item(numy, numx, 1) != 255 or blurs.item(numy, numx, 2) != 255:
                    checkx[numx-1] = True
                    break
        for numx in range(1, width):
            if checkx[numx-1] == False:
                blurs[height-1, numx] = [0, 0, 0]
                blurs[height-2, numx] = [0, 0, 0]
        # グラフの端を白塗り
        blurs = cv2.line(blurs, (0, 0), (0, height), (255, 255, 255), padding)
        blurs = cv2.line(blurs, (width, 0), (width, height),
                         (255, 255, 255), padding)
        # グラフの端に余白追加
        blurs = cv2.copyMakeBorder(
            blurs, 50, 50, 0, 0, cv2.BORDER_CONSTANT, value=[255, 255, 255])
        # 画像の2極化
        cv2.imwrite("autopeak_blurs_re.png", blurs)
        blurs = cv2.imread("autopeak_blurs_re.png")
        retval, im_bw = cv2.threshold(blurs, 254, 255, cv2.THRESH_BINARY)
        if isDebug == True:
            cv2.imwrite("blur_cut_re.png", im_bw)
        im_bw = cv2.cvtColor(im_bw, cv2.COLOR_BGR2GRAY)
        contours, hierarchy = cv2.findContours(
            im_bw, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
        # 余計な次元(余計な輪郭)削除 (NumContours, 1, NumPoints) -> (NumContours, NumPoints)
        contours = [np.squeeze(cnt, axis=1) for cnt in contours]
        conts = contours[0]
        # x座標範囲確認
        maxxc = 0
        minxc = 1000
        for cont in conts:
            if cont[0] >= maxxc:
                maxxc = cont[0]
            if cont[0] <= minxc:
                minxc = cont[0]

    # 一意に表示(Max)
    coodmax = {}
    for cont in conts:
        if cont[0] in coodmax:
            if cont[1] > coodmax[cont[0]]:
                coodmax[cont[0]] = cont[1]
        else:
            coodmax[cont[0]] = cont[1]
    # 一意に表示(Min)
    coodmin = {}
    for cont in conts:
        if cont[0] in coodmin:
            if cont[1] < coodmin[cont[0]]:
                coodmin[cont[0]] = cont[1]
        else:
            coodmin[cont[0]] = cont[1]
    # 一意に表示(Ave)
    cood = {}
    for coox in coodmax.keys():
        cood[coox] = (coodmax[coox]+coodmin[coox])//2

    if isDebug:
        print(cood)
        print("----------------------------")
        print(maxxc, minxc)
        if rangecheck == True:
            img = cv2.imread("blur_cut.png")
        else:
            img = cv2.imread("blur_cut_re.png")
        for coox in cood.keys():
            img[cood[coox], coox] = [0, 0, 255]
        cv2.imwrite("cood.png", img)

    return maxxc, minxc, cood


def modifypeak(k, cood):
    '''ピーク値の修正(y座標最大値をとるように)
    ・k: 実行時の入力値
    ・cood: グラフの座標
     return
    ・modifiedpx: 修正後x座標
    ・modifiedpy: 修正後y座標
    '''
    k = round(k*uns+zes[0]-xmin)
    try:
        prepeak = [k, cood[k]]
# 修正範囲
        modifiedpy = prepeak[1]
        for coox in range(prepeak[0]-trange, prepeak[0]+trange):
            if cood[coox] <= modifiedpy:
                modifiedpy = cood[coox]
                modifiedpx = coox
    except:
        modifiedpx = k
        modifiedpy = 1000000
        for i, j in cood.items():
            if modifiedpy > j:
                modifiedpy = j
    return modifiedpx, modifiedpy

# ピーク範囲の確定(トリミング画像内での)


def peakrange(maxxc, minxc, cood, modifiedpx, modifiedpy):
    '''ピーク範囲の確定
    ・maxxc: グラフのx軸の最大値
    ・minxc: グラフのx軸の最小値
    ・cood: グラフの座標
    ・modifiedpx: 修正後ピークx座標
    ・modifiedpy: 修正後ピークy座標
     return
    ・peakendx: ピークの終点x座標
    ・peakstax: ピークの起点x座標
    '''
    # ピーク下限調査(終点)
    for peakx in range(modifiedpx, maxxc-testx + 1):
        dist = cood[peakx + testx]-cood[peakx]
        thrhx = peakx + testx
        if dist < thrtest:
            break
    thrh = (modifiedpy + (equi-1)*cood[thrhx])/equi
    # ピーク範囲(終点)
    for peakx in range(modifiedpx, maxxc-searx + 1):
        dist = cood[peakx + searx]-cood[peakx]
        peakendx = peakx + searx
        if dist < heightthr and cood[peakx] > thrh:
            break
    # ピーク下限調査(始点)
    for peakx in range(modifiedpx, minxc + testx - 1):
        dist = cood[peakx - testx]-cood[peakx]
        thrhx = peakx - testx
        if dist < thrtest:
            break
    thrh = (modifiedpy + (equi-1)*cood[thrhx])/equi
    # ピーク範囲(始点)
    for peakx in range(modifiedpx, minxc + searx, -1):
        dist = cood[peakx-searx]-cood[peakx]
        peakstax = peakx - searx
        if dist < heightthr and cood[peakx] > thrh:
            break
    return peakendx, peakstax


def peakcodinate(name1, name2):
    '''ピークの矢印座標確認
    name1: ピーク矢印出現前スクリーンショットのpath
    name2: ピーク矢印出現後スクリーンショットのpath
     return
    x: 矢印のx座標
    y: 矢印のy座標
    '''
    ##ピーク座標の検出##
    # 画像１
    img_path_1 = r".\{}".format(name1)
    # 画像２
    img_path_2 = r".\{}".format(name2)
    # 画像読み込み
    img_1 = cv2.imread(img_path_1)
    img_1 = img_1[ymin:ymax, xmin:xmax]
    img_1 = cv2.cvtColor(img_1, cv2.COLOR_BGR2RGB)
    img_2 = cv2.imread(img_path_2)
    img_2 = img_2[ymin:ymax, xmin:xmax]
    img_2 = cv2.cvtColor(img_2, cv2.COLOR_BGR2RGB)
    # ピークチェックの色付け(2つの画像の差分を抽出)
    fgbg = cv2.createBackgroundSubtractorMOG2()
    fgmask = fgbg.apply(img_1)
    fgmask = fgbg.apply(img_2)
    cv2.imwrite(r".\autopeak_checked.png", fgmask)
    # 色の2極化
    im = cv2.imread('autopeak_checked.png')
    im_gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
    retval, im_bw = cv2.threshold(
        im_gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    # 輪郭の検出
    contours, hierarchy = cv2.findContours(
        im_bw, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    # 座標の検出(面積最大の差分についての重心から座標検出)
    areas = np.array(list(map(cv2.contourArea, contours)))
    max_idx = np.argmax(areas)
    max_area = areas[max_idx]
    result = cv2.moments(contours[max_idx])
    x = int(result["m10"]/result["m00"])
    y = int(result["m01"]/result["m00"])
    return x, y

def drawPeak(modifiedpx, peakendx, peakstax):
    '''ピークとベースラインを描く
    modifiedpx: 修正後ピークx座標
    peakendx: ピークの終わり座標(local)
    peakstax: ピークの始まり座標(local)
    '''
    #エラー処理はメインで
    ###ピークをつける###
    gui.click(positionPeakButton)
    gui.moveTo(modifiedpx + xmin,zes[1]-20)
    time.sleep(1)
    gui.doubleClick()
    #スクリーンショット
    gui.moveTo(positionOutRange)
    ImageGrab.grab().save("autopeak_before1.png")

    ###ベーススタート/エンド###
    gui.click(positionBaseBelowButton)
    time.sleep(0.5)
    #スクリーンショット
    gui.moveTo(positionOutRange)
    ImageGrab.grab().save("autopeak_after1.png")
    #ピーク座標矢印の検出
    x, y = peakcodinate("autopeak_before1.png","autopeak_after1.png")
    #カーソルの移動
    gui.moveTo(x + xmin, y + ymin)
    time.sleep(2)
    gui.dragTo(peakendx + xmin, y+ymin, 0.75)
    gui.moveTo(x + xmin, y + ymin - 3)
    gui.dragTo(peakstax + xmin, y+ymin, 0.75)
    #スクリーンショット
    gui.click(positionPeakButton)
    gui.moveTo(positionOutRange)
    ImageGrab.grab().save("autopeak_before2.png")

    ###ピークスタート/エンド###
    gui.click(positionBaseAboveButton)
    time.sleep(0.5)
    #スクリーンショット
    gui.moveTo(positionOutRange)
    ImageGrab.grab().save("autopeak_after2.png")
    #ピーク座標の検出#
    x, y = peakcodinate("autopeak_before2.png","autopeak_after2.png")
    #カーソルの移動
    gui.moveTo(x + xmin, y + ymin)
    time.sleep(2)
    gui.dragTo(peakendx + xmin, y+ymin, 0.75)
    gui.moveTo(x + xmin, y + ymin + 3)
    gui.dragTo(peakstax + xmin, y+ymin, 0.75)


def PasteImages(lst, photoNum_x, margin):
    '''目視確認画像の作成
    ・lst: 画像のpathのリスト
    ・photoNum_x: 横方向に並べる数
    ・margin: それぞれの画像ごとの間隔
     return
    ・canvas: 加工後の画像出力
    '''
    margin_pre = 10
    # 複数行ある場合
    if len(lst) > photoNum_x:
        # 画像サイズの確認
        image = Image.open(lst[0])
        width, height = image.size
        # 行数の計算
        photoNum_y = math.ceil(len(lst)/photoNum_x)
        # キャンバスの作成
        canvas = Image.new("RGB", (width*photoNum_x + margin*(photoNum_x + 1),
                           height*photoNum_y + margin*(photoNum_y + 1)), (189, 195, 201))
        for i in range(len(lst)):
            image = Image.open(lst[i])
            # 座標計算(0番目から)
            x_image = i % photoNum_x
            y_image = i//photoNum_x
            # ペースト
            # Recheckを強調
            if "checkphoto_1_F" in os.path.basename(lst[i]) or "checkphoto_2_F" in os.path.basename(lst[i]):
                canvas_pre = Image.new(
                    "RGB", (width + margin_pre, height + margin_pre), (255, 0, 0))
                canvas_pre.paste(image, (int(margin_pre/2), int(margin_pre/2)))
                image = canvas_pre
                paste_x = margin*(x_image+1) + width * \
                    x_image - int(margin_pre/2)
                paste_y = margin*(y_image+1) + height * \
                    y_image - int(margin_pre/2)
                canvas.paste(image, (paste_x, paste_y))
                title = os.path.basename(lst[i]).replace("checkphoto_1_F_", "").replace(
                    "checkphoto_2_F_", "").replace(".png", ".crm")
            else:
                canvas.paste(image, (margin*(x_image+1) + width *
                             x_image, margin*(y_image+1) + height * y_image))
                title = os.path.basename(lst[i]).replace("checkphoto_1_", "").replace(
                    "checkphoto_2_", "").replace(".png", ".crm")
            draw = ImageDraw.Draw(canvas)
            font = ImageFont.truetype('C:\windows\Fonts\meiryo.ttc', 27)
            draw.text((margin*(x_image+1) + width * x_image,
                       margin*(y_image+1) + height * (y_image+1)),
                      title,
                      font=font, fill="black")

    # 1行だけの場合(行サイズ調整)
    else:
        # 画像サイズの確認
        image = Image.open(lst[0])
        width, height = image.size
        # 行数の計算
        photoNum_y = 1
        photoNum_x = len(lst)
        # キャンバスの作成
        canvas = Image.new("RGB", (width*photoNum_x + margin*(photoNum_x + 1),
                           height*photoNum_y + margin*(photoNum_y + 1)), (189, 195, 201))
        for i in range(len(lst)):
            image = Image.open(lst[i])
            # 座標計算(0番目から)
            x_image = i
            y_image = 0
            # ペースト
            # Recheckを強調
            if "checkphoto_1_F" in os.path.basename(lst[i]) or "checkphoto_2_F" in os.path.basename(lst[i]):
                canvas_pre = Image.new(
                    "RGB", (width + margin_pre, height + margin_pre), (255, 0, 0))
                canvas_pre.paste(image, (int(margin_pre/2), int(margin_pre/2)))
                image = canvas_pre
                paste_x = margin*(x_image+1) + width * \
                    x_image - int(margin_pre/2)
                paste_y = margin*(y_image+1) + height * \
                    y_image - int(margin_pre/2)
                canvas.paste(image, (paste_x, paste_y))
                title = os.path.basename(lst[i]).replace("checkphoto_1_F_", "").replace(
                    "checkphoto_2_F_", "").replace(".png", ".crm")
            else:
                canvas.paste(image, (margin*(x_image+1) + width *
                             x_image, margin*(y_image+1) + height * y_image))
                title = os.path.basename(lst[i]).replace("checkphoto_1_", "").replace(
                    "checkphoto_2_", "").replace(".png", ".crm")
            draw = ImageDraw.Draw(canvas)
            font = ImageFont.truetype('C:\windows\Fonts\meiryo.ttc', 40)
            draw.text((margin*(x_image+1) + width * x_image,
                       margin*(y_image+1) + height * (y_image+1)),
                      title,
                      font=font, fill="black")
    return canvas


def ManageCheckPhotos(name, path):
    '''目視確認画像の作成
    ・name: 出力後目視画像の名前
    ・path: 目視確認画像の材料画像のkeyのpath
    '''
    margin = 96
    lst = glob.glob(path)
    photoNum_x = 4
    photoNum_y_max = 4

    # 貼り付け最大枚数以内の時
    if len(lst) <= photoNum_x * photoNum_y_max:
        im = PasteImages(lst, photoNum_x, margin)
        im.save(f"{name}.png")

    # 貼り付け最大枚数より多い時
    else:
        itr = len(lst) // (photoNum_x * photoNum_y_max)
        itr_con = photoNum_x * photoNum_y_max
        for i in range(itr+1):
            fin = min((i+1)*itr_con, len(lst))
            lst_image = lst[i*itr_con:fin]
            im = PasteImages(lst_image, photoNum_x, margin)
            im.save(f"{name}_{i}.png")


def CreateCheckPhotos(isCH4s=True):
    '''目視確認画像の作成をまとめたもの
    '''
    ManageCheckPhotos("Channel1", r".\checkphoto_1_*")
    if isCH4s:
        ManageCheckPhotos("Channel2", r".\checkphoto_2_*")

#---------------------------
#以下CRMtoEXCEL
#---------------------------


def copyfiles_start(_inputPath):
    '''外部フォルダからCRMtoEXCELフォルダにコピー'''
    path_CRMtoEXCEL = os.getcwd()
    files = glob.glob(r"{}/*.crm".format(_inputPath))
    for file in files:
        shutil.copy(file, f"{path_CRMtoEXCEL}")


def copyfiles_end(_outputPath):
    '''CRMtoEXCELフォルダから外部フォルダにコピー'''
    path_CRMtoEXCEL = os.getcwd()
    files = glob.glob(r"{}/*.crm".format(path_CRMtoEXCEL)) + \
        glob.glob(r"{}/Channel*.png".format(path_CRMtoEXCEL))
    os.makedirs(_outputPath, exist_ok=True)
    for file in files:
        try:
            shutil.move(file, f"{_outputPath}")
        except:
            os.remove(f"{_outputPath}/{os.path.basename(file)}")
            shutil.move(file, f"{_outputPath}")
    if os.path.exists(r"{}/Recheck".format(path_CRMtoEXCEL)):
        shutil.move("{}/Recheck".format(path_CRMtoEXCEL),
                    f"{_outputPath}/Recheck")


def maximizeCRM():
    '''グラフの最大化'''
    path = r".\images\maximizeButton.png"
    gui.click(gui.locateCenterOnScreen(path, confidence=0.7, grayscale=True))


def inputCheck():
    '''入力ボックスが消えてピークチェックできる状態になっているか確認
     return
    ・True: 入力ボックスが消えているのでOK
    ・False: 入力ボックスがまだ残っている, ダメ
    '''
    imagePath = r"./images/inputButton.png"
    check = gui.locateOnScreen(imagePath, confidence=0.7, grayscale=True)
    return not check


def Chanel2():
    '''Channel2に切り替え'''
    gui.click(positionChannel2)


def ExcelExport():
    '''エクセルに連続エクスポートを実行'''
    gui.click(positionFile)
    gui.click(positionContinuousExport)
    time.sleep(1)
    # エクスポートボタンを認識してクリック
    path = r".\images\exportButton.png"
    left, top, width, height = gui.locateOnScreen(
        path, confidence=0.6, grayscale=True)
    gui.click(left+width//4, top+height//2)


def ExcelFront():
    '''エクセルファイルを前面に'''
    try:
        CRMPro_interface = win32gui.FindWindow(None, "Book1 - Excel")
        time.sleep(4)
        win32gui.SetForegroundWindow(CRMPro_interface)
    except Exception as e:
        time.sleep(5)
        CRMPro_interface = win32gui.FindWindow(None, "Book1 - Excel")
        time.sleep(4)
        win32gui.SetForegroundWindow(CRMPro_interface)


def TempSave(name):
    '''ファイルを保存'''
    gui.hotkey("ctrl", "s")
    time.sleep(1.5)
    gui.write(name)
    time.sleep(2)
    gui.press("enter")


def exportcheck(_nom, _exname, _allFiles):
    '''エクスポートできなかったファイルを抽出
    ・_nom: シートナンバー
    ・_exname: 読み込むエクセルファイル名
    ・_allFiles: 処理したcrmファイル名(リスト)
     return
    ・Miss: 処理していないcrmファイル名(リスト)
    '''
    wb = px.load_workbook(f'{_exname}.xlsx')
    ws = wb[_nom]
    # ファイル最大行の取得
    maxr = ws.max_row
    sheet_range = ws["A2":f"A{maxr}"]
    lst = []
    # A列のA2から最後まで取得
    for row in sheet_range:
        for cell in row:
            # 該当セルの値取得
            cell_value = cell.value
            lst.append(cell_value)
    Miss = []
    for file in _allFiles:
        if file not in lst:
            Miss.append(file)
    return Miss


def excelarrange(_nom, _exName):
    '''エクセルに色付け
    _nom: シートナンバー
    _exName: 読み込むエクセルファイル名
    '''
    # ファイルを読み込んでシートを選択
    wb = px.load_workbook(f'{_exName}.xlsx')
    ws = wb[_nom]
    # 罫線の作成
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    # 罫線作成
    for rows in ws:
        for cell in rows:
            cell.border = border
    # 文字の色を変更
    # 背景色を変更
    maxr = ws.max_row
    for rows in ws['A1':'E1']:
        for cell in rows:
            cell.fill = PatternFill(
                patternType='solid', fgColor='D1FE7B', bgColor='D1FE7B')
            cell.font = Font(bold=True)
    for rows in ws["A2":f"A{maxr}"]:
        for cell in rows:
            cell.font = Font(bold=False)
    # 列幅調節
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        columnstr = px.utils.get_column_letter(column)

        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[columnstr].width = adjusted_width
    wb.save(f'{_exName}.xlsx')


def movefiles_end(_output_path):
    '''xlsxを移動'''
    path_CRMtoEXCEL = os.getcwd()
    files = glob.glob(r"{}/*.xlsx".format(path_CRMtoEXCEL))
    os.makedirs(_output_path, exist_ok=True)
    for file in files:
        try:
            shutil.move(file, f"{_output_path}")
        except:
            os.remove(f"{_output_path}/{os.path.basename(file)}")
            shutil.move(file, f"{_output_path}")


def pathCheck():
    '''読み込みエラーが発生しないようにチェック'''
    isPath = False
    if isPath:
        return
    imagePath = r"./images/checkPath.png"
    check = gui.locateOnScreen(imagePath, confidence=0.9, grayscale=True)
    if not check:
        raise NameError("ファイルを開く -> ファイルの場所が「CRMProAuto」じゃないのでは??")
    else:
        isPath = True
    return isPath
